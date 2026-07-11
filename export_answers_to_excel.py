"""
Runs a set of audit questions through answer_question.py's pipeline and exports
one Excel file with a row per question: the question itself, the sub-queries
generated for it, the synthesized answer, and the reference files it cited.

Usage: python export_answers_to_excel.py
Writes audit_qa_results.xlsx in the current directory.
"""
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from retrieval_agent import RetrievalAgent
from answer_question import run_question

QUESTIONS = [
    "How does the organization determine which products/devices are subject to design and development controls?",
    "What criteria are used to classify a product as a medical device under applicable regulatory jurisdictions?",
    "Are all devices, including accessories, software, variants, and families, appropriately identified and included?",
    "How are changes such as maintenance releases, configuration changes, or software updates evaluated to determine if design controls apply?",
    "Are legacy devices assessed for compliance with current design control requirements?",
    "How does the organization ensure global regulatory requirements are considered in determining design control applicability?",
    "Does the organization have documented design and development procedures compliant with applicable regulations?",
    "Are these procedures consistently applied to all identified devices?",
    "How does the organization ensure design controls are applied across the lifecycle for new, modified, and legacy devices?",
    "Are roles and responsibilities for design activities clearly defined?",
    "How is risk integrated into design and development activities?",
    "Is technical documentation established and maintained for each device?",
    "Does the documentation meet regulatory requirements for applicable markets?",
    "Is there a defined structure for technical documentation, such as DHF, STED, or Technical File?",
    "How is completeness and consistency of technical documentation ensured?",
    "How are updates and revisions to technical documentation controlled?",
    "Is traceability established from requirements to design to verification and validation?",
    "Does the technical documentation align with regulatory requirements of all applicable jurisdictions?",
    "Are country-specific requirements addressed, such as FDA DHF, EU Technical File, or Health Canada STED?",
    "How does the organization ensure updates to regulatory requirements are reflected in technical documentation?",
    "How is technical documentation controlled, including versioning, approvals, and access?",
    "Who has responsibility for maintaining technical documentation?",
    "Is documentation readily retrievable for audit and regulatory submission?",
    "Are document control processes followed?",
    "Is each device uniquely identified and linked to its technical documentation?",
    "How are product variants and configurations managed within technical documentation?",
    "Is traceability maintained across product families and versions?",
]

OUTPUT_XLSX = "audit_qa_results.xlsx"

COLUMNS = ["#", "Audit Question", "Sub-queries Used", "Answer", "Reference Files Used", "Chunks Used"]
COLUMN_WIDTHS = [4, 40, 40, 70, 40, 10]


def build_workbook(rows: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit QA Results"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, row in enumerate(rows, start=1):
        ws.append([
            i,
            row["question"],
            "\n".join(row["subqueries"]),
            row["answer"],
            "\n".join(row["documents_cited"]),
            row["chunks_used"],
        ])

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(COLUMNS)):
        for cell in row:
            cell.alignment = wrap

    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    return wb


if __name__ == "__main__":
    agent = RetrievalAgent(docs_folder="docs", csv_path="latest_revisions.csv")

    rows = []
    for i, question in enumerate(QUESTIONS, start=1):
        print(f"\n[{i}/{len(QUESTIONS)}] {question}")
        result = run_question(agent, question)
        rows.append({
            "question": question,
            "subqueries": result["subqueries"],
            "answer": result["answer"],
            "documents_cited": result["documents_cited"],
            "chunks_used": result["chunks_used"],
        })

    wb = build_workbook(rows)
    wb.save(OUTPUT_XLSX)
    print(f"\nSaved {len(rows)} rows to {OUTPUT_XLSX}")
